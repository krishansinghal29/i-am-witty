"""Build VGA exercise-analysis Excel from the extracted summaries.

Columns: Week | Section | Video | Idea in the Video | Can Form Exercise (bool)
         | Proposed / Related Exercise | Already Covered By | Notes / Rationale

Decision lens: riffy's exercises are SOLO verbal-wit drills (user generates a
clever line, AI evaluates). "Can Form Exercise" = TRUE only when the video
teaches a verbal line/response a solo user can practice generating. FALSE for
mindset/inner-game, physical/logistical mechanics, delivery/tonality, pure
strategy, or admin.
"""

import json
import os

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SUMMARIES = "/Users/krishansinghal/i-am-witty/design/extra/_summaries.json"
OUT = "/Users/krishansinghal/i-am-witty/design/extra/VGA_Exercise_Analysis.xlsx"

# path -> (can_form: bool, proposed_exercise, already_covered_by, note)
D = {
    "Read Me_transcript.json": (False, "", "", "Administrative read-me (empty transcript) — no teachable content."),

    # --- Week 0 / 01 Welcome ---
    "Week 0/01. Welcome/If You Have Approach Anxiety_transcript.json": (False, "", "", "Inner-game / approach-anxiety guidance — points to AA missions, no verbal line to drill."),
    "Week 0/01. Welcome/Preparing for the Program_transcript.json": (False, "", "", "Course-prep advice — administrative, no skill content."),
    "Week 0/01. Welcome/Welcome to the Program_transcript.json": (False, "", "", "Course intro/orientation — administrative."),

    # --- Week 0 / 02 OPEN-Close Model ---
    "Week 0/02. OPEN-Close Model/Close_transcript.json": (False, "", "", "Strategic concept (always aim to close). The ask is logistical; phrasing it well is better captured by the Rephrase/Phrasing drill."),
    "Week 0/02. OPEN-Close Model/Evaluation_transcript.json": (False, "", "questionAnswerTease", "Conceptual overview of qualification (ask/tease/lead). Drillable piece = the qualification tease (see Wk5 'Tease')."),
    "Week 0/02. OPEN-Close Model/How to Use the System_transcript.json": (False, "", "", "Framework overview (system vs verbal game) — no line to generate."),
    "Week 0/02. OPEN-Close Model/Narrative_transcript.json": (False, "", "", "Conceptual overview of narrative — strategic. Drillable parts live in Wk6 storytelling."),
    "Week 0/02. OPEN-Close Model/Open_transcript.json": (False, "", "", "Overview of opener types — conceptual; individual openers handled in Wk1."),
    "Week 0/02. OPEN-Close Model/Premise_transcript.json": (False, "", "pushPull", "Concept of conveying premise (negative/implied/push-pull). Drillable form = push-pull, already covered."),
    "Week 0/02. OPEN-Close Model/Steal My Lines_transcript.json": (False, "", "", "Resource (canned-lines PDF) — no skill to drill."),
    "Week 0/02. OPEN-Close Model/The Model_transcript.json": (False, "", "", "Open→Close framework overview — strategic, no line to generate."),

    # --- Week 0 / 03 Verbal Game Basics ---
    "Week 0/03. Verbal Game Basics/The Benefits of Verbal Game_transcript.json": (False, "", "", "Motivational/conceptual (why prepared lines help) — no drill."),
    "Week 0/03. Verbal Game Basics/What is Verbal Game_transcript.json": (False, "", "", "Defines verbal game — conceptual, no drill."),

    # --- Week 0 / 04 Eras of Game ---
    "Week 0/04. Eras of Game/Canned Game Era_transcript.json": (False, "", "", "History of pickup styles — conceptual."),
    "Week 0/04. Eras of Game/Introducing Optimized Game Era_transcript.json": (False, "", "", "History/philosophy of game styles — conceptual."),
    "Week 0/04. Eras of Game/Natural Game Era_transcript.json": (False, "", "", "History of pickup styles — conceptual."),
    "Week 0/04. Eras of Game/No Game Era_transcript.json": (False, "", "", "Mindset/history (game is a learnable skill) — conceptual."),

    # --- Week 0 / 05 Beyond Verbals ---
    "Week 0/05. Beyond Verbals/How to Deliver Lines_transcript.json": (False, "", "", "Delivery/tonality (how, not what) — can't drill as a text/line exercise."),
    "Week 0/05. Beyond Verbals/Inception_transcript.json": (False, "", "", "Strategic concept (make her chase you) — not a solo verbal drill."),
    "Week 0/05. Beyond Verbals/Non-Verbal Factors_transcript.json": (False, "", "", "Voice/eye-contact/pauses/inflection — delivery, not content."),
    "Week 0/05. Beyond Verbals/Non-Verbal Verbal Game_transcript.json": (False, "", "", "Gestures replacing words — non-verbal, not a line drill."),

    # --- Week 0 / 06 Key Concepts ---
    "Week 0/06. Key Concepts/Formats and Sequencing_transcript.json": (False, "", "", "Meta-concept (think in formats, Mad-Libs). Not itself a drill — but validates the whole format-based exercise approach."),
    "Week 0/06. Key Concepts/Invisible Escalation_transcript.json": (False, "", "", "Physical/verbal/logistical micro-escalation — about pacing, not a line to generate."),
    "Week 0/06. Key Concepts/Phrasing_transcript.json": (True, "Rephrase for Value", "", "Word choice changes perceived value (e.g. how you ask for a number). NEW drill: rewrite a needy/low-value line into a higher-value one. Sibling of Reframe & Connotation Swap."),

    # --- Week 0 / 07 Key Technical Themes ---
    "Week 0/07. Key Technical Themes/Baseline Vs. Plotline_transcript.json": (False, "", "", "Concept of fractionating baseline/plotline — strategic, no single line."),
    "Week 0/07. Key Technical Themes/Building a Narrative Into the Interaction_transcript.json": (True, "Instant Narrative / Role", "yesAnd", "Assign a playful instant dynamic (e.g. 'you're my runaway ex') and banter from it. NEW creative role/premise drill; cousin of yesAnd and the roleplay engine."),
    "Week 0/07. Key Technical Themes/Calibration _ Blueprint_transcript.json": (False, "", "", "Concept (tailor to the individual) — strategic, no line."),
    "Week 0/07. Key Technical Themes/How to Be Authentic_transcript.json": (False, "", "", "Mindset (stay authentic while using lines) — no drill."),

    # --- Week 0 / 09 Approach Anxiety ---
    "Week 0/09. Approach Anxiety/Approach Excitement_transcript.json": (False, "", "", "Approach-anxiety reframe (fear = excitement) — inner game, not verbal."),
    "Week 0/09. Approach Anxiety/Emotional Leverage Solution_transcript.json": (False, "", "", "Approach-anxiety mindset (fear of self-rejection) — inner game."),
    "Week 0/09. Approach Anxiety/Introduction_transcript.json": (False, "", "", "Intro to approach anxiety — inner game."),
    "Week 0/09. Approach Anxiety/Physiology Solution_transcript.json": (False, "", "", "Physical/state technique (posture, dancing, shouting) — not verbal."),
    "Week 0/09. Approach Anxiety/Reframe Solution_transcript.json": (False, "", "", "Approach-anxiety reframe (just say hi) — inner game."),
    "Week 0/09. Approach Anxiety/Technical Solutions_transcript.json": (False, "", "", "Overview of AA solution methods — inner game."),
    "Week 0/09. Approach Anxiety/Trick Your Brain Solution_transcript.json": (False, "", "", "Approach-anxiety trick (gun-to-the-head start) — inner game."),

    # --- Week 0 / 10-11 Missions ---
    "Week 0/10. Beginner Missions/Ask for the Number_transcript.json": (False, "", "", "Field mission (go ask for the number) — not a solo verbal drill."),
    "Week 0/10. Beginner Missions/Assume the Burden_transcript.json": (False, "", "", "Field mission (carry the conversation 10s) — not a solo verbal drill."),
    "Week 0/11. Advanced Missions/Groups_transcript.json": (False, "", "", "Field mission (approach groups) — physical scenario practice."),
    "Week 0/11. Advanced Missions/Headphones_transcript.json": (False, "", "", "Field mission (approach girls in headphones) — physical scenario."),
    "Week 0/11. Advanced Missions/Introduction_transcript.json": (False, "", "", "Intro to the AA challenge missions — no content."),
    "Week 0/11. Advanced Missions/Man to Woman_transcript.json": (False, "", "", "Field mission (break the customer frame) — physical scenario."),
    "Week 0/11. Advanced Missions/Other Guys_transcript.json": (False, "", "", "Field mission (approach when other men present) — physical scenario."),
    "Week 0/11. Advanced Missions/Public Eye_transcript.json": (False, "", "", "Field mission (approach in public) — physical scenario."),
    "Week 0/11. Advanced Missions/Social Pressure_transcript.json": (False, "", "", "Field mission (approach in unusual venues) — physical scenario."),
    "Week 0/11. Advanced Missions/Switch Styles_transcript.json": (False, "", "", "Field practice switching direct/indirect — mechanics/mindset, not a line drill."),
    "Week 0/11. Advanced Missions/Walking_transcript.json": (False, "", "", "Field mission (approach girls in motion) — physical scenario."),

    # --- Week 1 / Approach Anxiety ---
    "Week 1/1 - Approach Anxiety/1- Introduction to Approach Anxiety_transcript.json": (False, "", "", "Intro to approach anxiety — inner game."),
    "Week 1/1 - Approach Anxiety/10 - Shock and Awe_transcript.json": (False, "", "", "AA desensitization method (bold approaches) — inner game."),
    "Week 1/1 - Approach Anxiety/2- Causes_transcript.json": (False, "", "", "Causes of approach anxiety (evolution/ego) — inner game."),
    "Week 1/1 - Approach Anxiety/3- Mentalities — Force Your Brain_transcript.json": (False, "", "", "AA mentality (just take the physical action) — inner game."),
    "Week 1/1 - Approach Anxiety/4- Mentalities — What is Rejection#_transcript.json": (False, "", "", "AA mentality (rejection is contextual) — inner game."),
    "Week 1/1 - Approach Anxiety/5- Mentalities — Simulation_transcript.json": (False, "", "", "AA mentality (treat it like a simulation) — inner game."),
    "Week 1/1 - Approach Anxiety/6 - Mentalities — Numbers Game_transcript.json": (False, "", "", "AA mentality (it's a numbers game) — inner game."),
    "Week 1/1 - Approach Anxiety/7 - Mentalities — Value Giving_transcript.json": (False, "", "", "AA mentality (you give value) — inner game."),
    "Week 1/1 - Approach Anxiety/8 - Mentalities — Winning_transcript.json": (False, "", "", "AA mentality (every approach is a win) — inner game."),
    "Week 1/1 - Approach Anxiety/9 - Progressive Desensitization_transcript.json": (False, "", "", "AA method (gradual exposure) — inner game."),

    # --- Week 1 / Approach Angles ---
    "Week 1/2 - Approach Angles/1 - Introduction to Approach Angles_transcript.json": (False, "", "", "Physical approach mechanics intro (body language/positioning) — not verbal."),
    "Week 1/2 - Approach Angles/2 - Seated_transcript.json": (False, "", "", "Physical mechanics (approaching a seated girl) — not verbal."),
    "Week 1/2 - Approach Angles/3 - Standing_transcript.json": (False, "", "", "Physical mechanics (standing approach, venue change) — not verbal."),
    "Week 1/2 - Approach Angles/4 - Moving_transcript.json": (False, "", "", "Physical mechanics (stopping a girl in motion) — not verbal."),
    "Week 1/2 - Approach Angles/5 - Behind and Front_transcript.json": (False, "", "", "Physical mechanics (approach from behind/front) — not verbal."),
    "Week 1/2 - Approach Angles/6 - Loud Environment_transcript.json": (False, "", "", "Physical mechanics (volume/closeness in clubs) — delivery, not content."),
    "Week 1/2 - Approach Angles/7 - Groups_transcript.json": (False, "", "", "Physical mechanics (managing group + target) — not a solo line drill."),

    # --- Week 1 / Openers ---
    "Week 1/3 - Openers/1 - Introduction to Openers_transcript.json": (False, "", "", "Overview of openers — conceptual."),
    "Week 1/3 - Openers/2 - Direct Open_transcript.json": (False, "", "", "Direct compliment opener — about courage/delivery; minimal wit to generate."),
    "Week 1/3 - Openers/3 - Indirect Open_transcript.json": (False, "", "", "Concept (indirect vs direct pacing) — no specific line."),
    "Week 1/3 - Openers/4 -  Push-Pull Open_transcript.json": (True, "(Push-Pull)", "pushPull", "Push-pull opener — already covered by the pushPull exercise."),
    "Week 1/3 - Openers/5 - Situational Open_transcript.json": (False, "", "", "Comment on the environment — low-wit/safe; stronger sibling is the observational/tease opener."),
    "Week 1/3 - Openers/6 - Observational Open_transcript.json": (True, "(First Unusual Thing)", "firstUnusualThing", "Tease about something in the moment + exaggerate — already covered by firstUnusualThing."),
    "Week 1/3 - Openers/7 - Non-Verbal Open_transcript.json": (False, "", "", "Winks/gestures/beckoning — non-verbal."),
    "Week 1/3 - Openers/8 - Opinion Open_transcript.json": (False, "", "", "Opinion opener — leans on a hook/story; better served by the Storytelling drill than its own exercise."),

    # --- Week 2 / Attraction Fundamentals ---
    "Week 2/1 - Attraction Fundamentals/1- Introduction to Attraction_transcript.json": (False, "", "", "Intro/overview of attraction — conceptual."),
    "Week 2/1 - Attraction Fundamentals/10- Misinterpret_transcript.json": (True, "(Misinterpretation)", "misinterpretation / sexualMisinterpretation", "The misinterpretation drill itself (take a sentence with I/you/we, read it as a compliment/advance) — already covered."),
    "Week 2/1 - Attraction Fundamentals/11- Shit Test_transcript.json": (True, "(Shit Test)", "shitTest", "Passing shit tests via misinterpret / agree & exaggerate — already covered by shitTest."),
    "Week 2/1 - Attraction Fundamentals/12- Giving Shit Tests_transcript.json": (True, "Give a Shit Test", "", "Flip side of shitTest: craft a playful challenge/tease to GIVE her. NEW drill; complements shitTest (passing) and pushPull."),
    "Week 2/1 - Attraction Fundamentals/2- Causes of Attraction_transcript.json": (False, "", "", "Evo-psych theory (status, social proof, non-neediness) — conceptual."),
    "Week 2/1 - Attraction Fundamentals/3- Premise_transcript.json": (False, "", "pushPull", "Premise vs intent concept — drillable form (negative phrasing/push-pull) already covered."),
    "Week 2/1 - Attraction Fundamentals/4- Don't Supplicate_transcript.json": (False, "", "", "Mindset (don't beg/change yourself) — inner game."),
    "Week 2/1 - Attraction Fundamentals/5- You are Enough_transcript.json": (False, "", "", "Mindset (you are enough) — inner game."),
    "Week 2/1 - Attraction Fundamentals/6- Attraction Isn't a Choice_transcript.json": (False, "", "", "Concept (attraction is instinctive) — conceptual."),
    "Week 2/1 - Attraction Fundamentals/7- Act as if_transcript.json": (False, "", "", "Mindset (act as your aspirational self) — inner game."),
    "Week 2/1 - Attraction Fundamentals/8- Sexualize_transcript.json": (True, "That's What She Said", "sexWithMeIsLike / sexualMisinterpretation", "Suggests 'that's what she said' and 'sex with me is like' drills. 'Sex with me is like' = sexWithMeIsLike (covered). 'That's What She Said' (spot/insert the innuendo) is a NEW drill, sibling of sexualMisinterpretation."),
    "Week 2/1 - Attraction Fundamentals/9- Play to Win_transcript.json": (False, "", "", "Mindset (take risks, make a strong impression) — inner game."),

    # --- Week 3 / Attraction Techniques ---
    "Week 3/1 - Attraction Techniques/01 -  Introduction to Attraction Techniques_transcript.json": (False, "", "", "Intro/overview — conceptual."),
    "Week 3/1 - Attraction Techniques/02 - Push-Pull_transcript.json": (True, "(Push-Pull)", "pushPull", "Core push-pull technique — already covered by pushPull."),
    "Week 3/1 - Attraction Techniques/03 - Cold Read_transcript.json": (True, "Cold Read", "", "Playfully TELL someone something about themselves (funny/exaggerated guess) instead of asking. STRONG new drill: given a cue/photo/trait, generate a witty cold read."),
    "Week 3/1 - Attraction Techniques/04 - Exaggeration_transcript.json": (True, "Agree & Exaggerate", "yesAnd", "Exaggeration humor + 'agree & exaggerate' to absurd extremes. NEW drill: take a premise/accusation and amplify it to a funny extreme. Distinct from yesAnd (punch = absurd amplification)."),
    "Week 3/1 - Attraction Techniques/05 - Cocky Frame_transcript.json": (True, "Cocky Frame", "shitTest", "Adopt a confident cocky-funny frame ('I came over to amuse myself'). NEW drill: open/respond with a playfully arrogant, high-value line."),
    "Week 3/1 - Attraction Techniques/06 - Open Loop_transcript.json": (True, "Open Loop", "", "Create intrigue by leaving a thread unfinished ('You remind me of... actually never mind'). NEW drill: turn a statement into an intriguing cliffhanger."),
    "Week 3/1 - Attraction Techniques/07 - Disqualify_transcript.json": (True, "Disqualifier", "pushPull", "Playful disqualification ('you're not my type, we'd never work because...'). NEW drill: generate a charming disqualifier that makes her requalify."),
    "Week 3/1 - Attraction Techniques/08 - Intensity_transcript.json": (False, "", "", "A MODIFIER (how explicit/hard a tease is) — best as an intensity setting inside the tease drills, not its own exercise."),
    "Week 3/1 - Attraction Techniques/09 - Delivery_transcript.json": (False, "", "", "Tonality/delivery of a tease — how, not what; can't drill as a line."),
    "Week 3/1 - Attraction Techniques/10 - Softener_transcript.json": (False, "", "", "Adding qualifiers to soften a tease — a modifier usable inside tease drills; thin as a standalone exercise."),
    "Week 3/1 - Attraction Techniques/11 - Association_transcript.json": (False, "", "", "Teasing via association/3rd party — a tease variant/modifier rather than its own drill."),
    "Week 3/1 - Attraction Techniques/12 -  Finality_transcript.json": (False, "", "", "Holding tension/silence after a tease — timing/delivery, not content."),
    "Week 3/1 - Attraction Techniques/13 - Moving Forward_transcript.json": (False, "", "", "Concept: transition from attraction to connection — strategic, no line."),

    # --- Week 4 / Frame Control ---
    "Week 4/1 - Frame Control/1 - Introduction to Frame Control_transcript.json": (False, "", "", "Intro/overview of frame control — conceptual."),
    "Week 4/1 - Frame Control/10 - Convinced Vs. Seduced_transcript.json": (False, "", "", "Ethics/concept (persuade vs seduce vs force) — no drill."),
    "Week 4/1 - Frame Control/11 -  Seeding & Babystepping_transcript.json": (False, "", "", "Incremental-commitment tactic — logistical/strategic, not a line."),
    "Week 4/1 - Frame Control/12 - Preframing_transcript.json": (False, "", "", "Set the frame before an objection — strategic; complex to drill (see Reframe)."),
    "Week 4/1 - Frame Control/13 - Situational Opportunity_transcript.json": (False, "", "", "Opportunism/timing in venues — logistical."),
    "Week 4/1 - Frame Control/14 -Social Proof_transcript.json": (False, "", "", "Wing/social-proof tactics — logistical, not a solo line."),
    "Week 4/1 - Frame Control/15 - Misinterpretation_transcript.json": (True, "(Misinterpretation)", "misinterpretation / shitTest", "Reframe platonic as sexual / pass shit tests via misinterpretation — already covered."),
    "Week 4/1 - Frame Control/16 - Recap_transcript.json": (False, "", "", "Recap of frame control — no content."),
    "Week 4/1 - Frame Control/17 - Missions for Week 4_transcript.json": (True, "Reframe (Positive Spin)", "ifByXYouMeanY", "Mission: reframe everything positively; assert a frame fast. NEW broad 'Reframe' drill (spin a negative situation/line positively). ifByXYouMeanY covers reframing criticism into status; this is the general version."),
    "Week 4/1 - Frame Control/2 - What is Frame Control#_transcript.json": (False, "", "", "Defines 'frame' (NLP roots) — conceptual."),
    "Week 4/1 - Frame Control/3 -  Choosing Positive Frames_transcript.json": (False, "", "ifByXYouMeanY", "Concept behind reframing positively — drillable version = the Reframe exercise (Wk4 missions)."),
    "Week 4/1 - Frame Control/4 - The Four Levels of Learning Frame Control_transcript.json": (False, "", "", "Learning-stages concept — no drill."),
    "Week 4/1 - Frame Control/5 - Cialdini Concepts_transcript.json": (False, "", "", "Cialdini's 6 influence principles — theory, not a line drill."),
    "Week 4/1 - Frame Control/6 - Denotation Vs. Connotation_transcript.json": (True, "Connotation Swap", "", "Same denotation, different connotation ('player' vs 'outgoing social guy'). NEW drill: restate something in a more attractive/charming connotation. Sibling of Reframe/Rephrase."),
    "Week 4/1 - Frame Control/7 - Perceived Vs. Expected_transcript.json": (False, "", "", "Framing a question to evoke behavior — strategic, no single line."),
    "Week 4/1 - Frame Control/8 - Social Pressure_transcript.json": (False, "", "", "Non-verbal influence/pressure — delivery/mechanics."),
    "Week 4/1 - Frame Control/9 - Authority_transcript.json": (False, "", "", "Projecting authority via passion/knowledge — mindset/concept."),

    # --- Week 5 / Evaluation ---
    "Week 5/1 - Evaluation/1 - Introduction to Evaluation_transcript.json": (False, "", "", "Intro/overview of evaluation — conceptual."),
    "Week 5/1 - Evaluation/2 - Why Evaluation Matters_transcript.json": (False, "", "", "Why qualification matters — conceptual."),
    "Week 5/2 - How to Evaluate/1 - Ask_transcript.json": (False, "", "questionAnswerTease", "Direct qualification question — a single canned ask; the charming, drillable version is the Qualification Tease."),
    "Week 5/2 - How to Evaluate/2 - Tease_transcript.json": (True, "Qualification Tease", "questionAnswerTease", "Tease/challenge her into qualifying herself. NEW drill: give a playful challenge that makes her rise to it. Overlaps questionAnswerTease; could extend it."),
    "Week 5/2 - How to Evaluate/3 - Lead_transcript.json": (False, "", "vibing", "Share a passion then pause to invite her in — a conversational lead; closest to vibing, hard to drill as one line."),
    "Week 5/2 - How to Evaluate/4 - Have Value_transcript.json": (False, "", "", "Be high-value so she qualifies herself — concept/mindset."),
    "Week 5/2 - How to Evaluate/5 - Qualification_transcript.json": (False, "", "", "Let her 'win', balance +/- — strategic concept."),
    "Week 5/3 - Missions/Missions for Week 5_transcript.json": (False, "", "", "Homework: list non-physical traits you value + screen for them — introspective/field, not a verbal drill."),

    # --- Week 6 / Narrative ---
    "Week 6/1 - Narrative/1 - What is Narrative and Why is it Last_transcript.json": (False, "", "", "Defines narrative — conceptual."),
    "Week 6/1 - Narrative/10 - Plotline as Narrative# Selectivity_transcript.json": (False, "", "", "Concept (be selective to raise value) — strategic."),
    "Week 6/1 - Narrative/11- Plotline as Narrative# Lack of Agenda_transcript.json": (False, "", "", "Concept (appear agenda-less) — strategic."),
    "Week 6/1 - Narrative/12- Plotline as Narrative# Knowing the Dance_transcript.json": (False, "", "", "Concept (calibrated escalation by experience) — strategic."),
    "Week 6/1 - Narrative/13 - Plotline as Narrative# Social Proof_transcript.json": (False, "", "", "Concept (social proof > direct game) — strategic."),
    "Week 6/1 - Narrative/14 - Baseline as Narrative# Your Lifestyle_transcript.json": (False, "", "", "Concept (convey lifestyle via passion, not as a provider) — strategic; Storytelling drill covers the doable part."),
    "Week 6/1 - Narrative/15 - Baseline as Narrative# Your Archetype_transcript.json": (False, "", "", "Concept (choose your own archetype) — identity/strategic."),
    "Week 6/1 - Narrative/16 - Baseline as Narrative# Your Sexual Story_transcript.json": (False, "", "", "Framing your sexual history — niche/introspective, not a wit drill."),
    "Week 6/1 - Narrative/17 - Baseline as Narrative# Teaching and Leading_transcript.json": (False, "", "", "Conversational role (teacher/leader) concept — strategic."),
    "Week 6/1 - Narrative/18 -  Narrative - Final Thoughts_transcript.json": (False, "", "", "Concept (narrative dominates long-term) — strategic."),
    "Week 6/1 - Narrative/19 -  Blueprint Intro_transcript.json": (False, "", "", "Intro to the Blueprint concept — conceptual."),
    "Week 6/1 - Narrative/2 - Narrative Technique 1 - General Storytelling_transcript.json": (True, "Storytelling (3-Act)", "", "Tell a story with beginning/middle/end; practice by spinning a mini-story about a random object. NEW drill: given a prompt/object, tell an engaging 3-beat micro-story. (Longer-form than current drills but high value.)"),
    "Week 6/1 - Narrative/3 - Narrative Technique 2 - Purposeful Storytelling_transcript.json": (True, "Purposeful Story", "", "Choose/shape a story to convey a desired trait. NEW drill (variant of Storytelling): tell a story that demonstrates a given trait. Pairs with Show Don't Tell."),
    "Week 6/1 - Narrative/4 - Narrative Technique 3 - Objective Correlative (Show Don't Tell)_transcript.json": (True, "Show Don't Tell", "", "Convey a quality through events/details, not by stating it. NEW drill: imply a trait via a concrete detail/story instead of claiming it."),
    "Week 6/1 - Narrative/5 - Narrative Technique 4 - Inception_transcript.json": (False, "", "", "Guide them to a conclusion via questions — strategic/multi-turn, hard to drill solo."),
    "Week 6/1 - Narrative/6 - Narrative Technique 5 - Preframing & Reframing_transcript.json": (False, "", "ifByXYouMeanY", "Meta-frame pre/reframing — strategic; reframing drill covered by Reframe & ifByXYouMeanY."),
    "Week 6/1 - Narrative/7 - Narrative of YOU_transcript.json": (False, "", "", "Self-presentation / prep your story — introspective prep, not a generative drill."),
    "Week 6/1 - Narrative/8 - Narrative of WE_transcript.json": (False, "", "", "Concept (you-as-a-couple narrative) — strategic."),
    "Week 6/1 - Narrative/9 - Baseline vs. Plotline Intro_transcript.json": (False, "", "", "Concept (baseline vs plotline) — strategic."),

    # --- Week 7 / Sexual Escalation (all physical/logistical) ---
    "Week 7/1 - Sexual Escalation/1 - Introduction to Escalation and Verbal Game_transcript.json": (False, "", "", "Intro to escalation (verbal+physical+venue must move together) — conceptual."),
    "Week 7/1 - Sexual Escalation/10 - How to Kiss - Introduction_transcript.json": (False, "", "", "Physical (kissing as non-verbal message) — not verbal."),
    "Week 7/1 - Sexual Escalation/11 - How to Kiss - Decision_transcript.json": (False, "", "", "Physical/mindset (decide to go for the kiss) — not verbal."),
    "Week 7/1 - Sexual Escalation/12 - How to Kiss - Phase Shift_transcript.json": (False, "", "sexualMisinterpretation", "Shifting to intimacy — mostly proximity/eye-contact/delivery; verbal sexual play already covered by sexualMisinterpretation/sexWithMeIsLike."),
    "Week 7/1 - Sexual Escalation/13 - How to Kiss - Tactics_transcript.json": (False, "", "", "Physical kissing tactics (gazing, escalation ladder) — not verbal."),
    "Week 7/1 - Sexual Escalation/14 - When Not to Escalate - Venue_transcript.json": (False, "", "", "Logistical (when/where NOT to escalate) — strategic."),
    "Week 7/1 - Sexual Escalation/15 - When Not to Escalate - LMR - Point of No Return_transcript.json": (False, "", "", "Physical/logistical (point of no return, LMR) — not verbal."),
    "Week 7/1 - Sexual Escalation/16 - When Not to Escalate - Transitions_transcript.json": (False, "", "", "Logistical (don't escalate during transitions) — strategic."),
    "Week 7/1 - Sexual Escalation/17 - When Not to Escalate - Pulling too Early_transcript.json": (False, "", "", "Logistical (don't pull too early) — strategic."),
    "Week 7/1 - Sexual Escalation/18 -  Sex - Relationship Blueprint_transcript.json": (False, "", "", "Concept (read her sexual/relationship blueprint) — strategic/reading."),
    "Week 7/1 - Sexual Escalation/2 - Escalation Ladder - Moves-Lines_transcript.json": (False, "", "", "Concept of staged verbal+physical escalation — strategic; per-stage lines are covered by the individual wit drills."),
    "Week 7/1 - Sexual Escalation/3 - Escalation Ladder - 2 Forward 1 Back_transcript.json": (False, "", "", "Pacing concept (fractionation) — strategic."),
    "Week 7/1 - Sexual Escalation/4 - Escalation Ladder - 80 Method_transcript.json": (False, "", "", "Pacing concept (stop at 80%) — strategic."),
    "Week 7/1 - Sexual Escalation/5 - Introduction to Venue Escalation_transcript.json": (False, "", "", "Logistical (venues carry social rules) — strategic."),
    "Week 7/1 - Sexual Escalation/6 - Venue Escalation - Meet Venue_transcript.json": (False, "", "", "Logistical (meet venue → move on) — strategic."),
    "Week 7/1 - Sexual Escalation/7 -  Venue Escalation - Isolate Venue_transcript.json": (False, "", "", "Logistical (isolation venue) — strategic."),
    "Week 7/1 - Sexual Escalation/8 - Venue Escalation - Pull - Date Venue_transcript.json": (False, "", "", "Logistical (date venue; mentions 'never have I ever'/'truth or dare' as games, not a wit drill)."),
    "Week 7/1 - Sexual Escalation/9 - Venue Escalation - Intimacy Venue_transcript.json": (False, "", "", "Logistical/physical (intimacy venue, gradual escalation) — not verbal."),

    # --- Week 8 / Blueprint ---
    "Week 8/1 - Blueprint/1 - What is Blueprint_transcript.json": (False, "", "", "Concept (calibrate to the individual; poker metaphor) — strategic/reading."),
    "Week 8/1 - Blueprint/10 - Comfort Blueprint_transcript.json": (False, "", "", "Concept (comfort varies per person) — reading/calibration."),
    "Week 8/1 - Blueprint/11 - Cultural Factors_transcript.json": (False, "", "", "Concept (culture shapes her) — reading."),
    "Week 8/1 - Blueprint/12 - Affects Blueprint_transcript.json": (False, "", "", "Concept (her past with men shapes her) — reading."),
    "Week 8/1 - Blueprint/13 - Eliciting Values Routine_transcript.json": (True, "Deepening Question", "vibing", "Ask about her passions then dig into the FEELING behind them. NEW connection drill: turn a surface answer into an emotional 'why' question. Overlaps vibing (attunement) but narrower & teachable."),
    "Week 8/1 - Blueprint/14 - Ends values over Means values_transcript.json": (False, "", "", "Concept (end vs means values) — strategic; informs Deepening Question."),
    "Week 8/1 - Blueprint/15 - When & How to Feed Values Back to Her_transcript.json": (False, "", "", "Timing concept (don't parrot her interests) — strategic."),
    "Week 8/1 - Blueprint/16 - Sexual Blueprint_transcript.json": (False, "", "", "Concept (read her sexual blueprint) — reading."),
    "Week 8/1 - Blueprint/17 - General Blueprint vs Blueprint for the Night_transcript.json": (False, "", "", "Concept (state/context shifts her) — reading."),
    "Week 8/1 - Blueprint/2 - Optional Poker Lesson for Deeper Understanding_transcript.json": (False, "", "", "Poker analogy for calibration — conceptual."),
    "Week 8/1 - Blueprint/3 - How Blueprint Adds to Solid Game_transcript.json": (False, "", "", "Concept (blueprint complements fundamentals) — strategic."),
    "Week 8/1 - Blueprint/4 - Do not Supplicate_transcript.json": (False, "", "", "Mindset (don't supplicate; adjust only within a comfortable range) — inner game."),
    "Week 8/1 - Blueprint/5 - Facets of You_transcript.json": (False, "", "", "Concept (show different facets per context) — identity/mindset."),
    "Week 8/1 - Blueprint/6 - Emotional Matches Over Logical Ones_transcript.json": (False, "", "", "Concept (emotional > logical similarity) — strategic; informs Deepening Question."),
    "Week 8/1 - Blueprint/7 - Whelmed_transcript.json": (False, "", "", "Concept (be 'whelmed' — neither over- nor under-invested) — mindset/calibration."),
    "Week 8/1 - Blueprint/8 - Guess and Check_transcript.json": (False, "", "", "Concept (guess her blueprint, adjust) — strategic/reading."),
    "Week 8/1 - Blueprint/9 - Value Blueprint & Value to Comfort Ratio_transcript.json": (False, "", "", "Concept (value-to-comfort ratio) — reading/calibration."),

    # --- Week 8 / Archetypes (great roleplay-persona fuel, not wit drills) ---
    "Week 8/2 - Archetypes/1 - The Princess_transcript.json": (False, "", "", "Target persona ('the Princess') + how to engage her. Not a standalone wit drill, but HIGH-VALUE input for roleplay characters/personas (feeds the roleplay character generator)."),
    "Week 8/2 - Archetypes/2 - Smart-Good Girl_transcript.json": (False, "", "", "Target persona ('smart good girl') + engagement strategy. Not a wit drill, but HIGH-VALUE input for roleplay personas."),
    "Week 8/2 - Archetypes/3 - First Generation American_transcript.json": (False, "", "", "Target persona ('first-gen American') + engagement strategy. Not a wit drill, but HIGH-VALUE input for roleplay personas."),
    "Week 8/2 - Archetypes/4 - The Entrepreneur_transcript.json": (False, "", "", "Target persona ('the entrepreneur') + engagement strategy. Not a wit drill, but HIGH-VALUE input for roleplay personas."),
}


def parse_path(path):
    parts = path.replace("_transcript.json", "").split("/")
    if len(parts) == 1:
        return "(root)", "", parts[0]
    if len(parts) == 2:
        return parts[0], "", parts[1]
    return parts[0], parts[1], "/".join(parts[2:])


def main():
    with open(SUMMARIES) as f:
        rows = json.load(f)

    missing = [r["path"] for r in rows if r["path"] not in D]
    if missing:
        raise SystemExit(f"Missing decisions for {len(missing)} paths:\n" + "\n".join(missing))

    wb = openpyxl.Workbook()

    # ---- Sheet 1: full catalog ----
    ws = wb.active
    ws.title = "VGA Video Analysis"
    headers = [
        "Week", "Section", "Video", "Idea in the Video",
        "Can Form Exercise", "Proposed / Related Exercise",
        "Already Covered By", "Notes / Rationale",
    ]
    ws.append(headers)

    for r in rows:
        week, section, video = parse_path(r["path"])
        can, ex_name, covered, note = D[r["path"]]
        idea = r["abstract"] or "(No transcript content — administrative file.)"
        ws.append([week, section, video, idea, bool(can), ex_name, covered, note])

    # styling
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(bold=True, color="FFFFFF")
    yes_fill = PatternFill("solid", fgColor="C6EFCE")
    for c in ws[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    widths = [9, 26, 42, 70, 16, 26, 26, 70]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
        if row[4].value is True:  # highlight TRUE rows green
            row[4].fill = yes_fill
            row[4].font = Font(bold=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{ws.max_row}"

    # ---- Sheet 2: proposed NEW exercises summary ----
    ws2 = wb.create_sheet("New Exercise Ideas")
    ws2.append(["#", "Proposed Exercise", "Source Video(s)", "What the user practices", "Relation to existing"])
    new_ex = [
        ("Cold Read", "Wk3 / 03 Cold Read",
         "Given a person-cue (photo/trait/job), make a playful, exaggerated guess about who they are instead of asking questions.",
         "New. Adjacent to firstUnusualThing (observation) but it's a confident assertion, not a tease about a thing."),
        ("Agree & Exaggerate", "Wk3 / 04 Exaggeration",
         "Take an accusation/premise and amplify it to an absurd, funny extreme ('yep, I'm exactly that, and worse...').",
         "New. Distinct from yesAnd — the payoff is absurd amplification; also the canonical shit-test defense."),
        ("Disqualifier", "Wk3 / 07 Disqualify",
         "Generate a charming 'you're not my type because...' that makes her want to requalify.",
         "New. Cousin of pushPull (the 'push' taken to a playful rejection)."),
        ("Open Loop", "Wk3 / 06 Open Loop",
         "Turn a statement into an intriguing cliffhanger that makes the other person chase the rest.",
         "New. Pure intrigue mechanic; not covered."),
        ("Cocky Frame", "Wk3 / 05 Cocky Frame",
         "Open/respond with a playfully arrogant, high-value line that flips who's impressing whom.",
         "New. Overlaps shitTest tone; different goal (set a frame, not pass one)."),
        ("Give a Shit Test", "Wk2 / 12 Giving Shit Tests",
         "Craft a playful challenge/tease to GIVE her (the active side of frame control).",
         "New. The mirror of shitTest, which only trains passing them."),
        ("That's What She Said", "Wk2 / 08 Sexualize",
         "Spot the line in a conversation where an innuendo lands, and deliver it.",
         "New, but close sibling of sexualMisinterpretation; could be a mode of it."),
        ("Reframe (Positive Spin)", "Wk4 / 17 Missions; Wk4 / 03",
         "Take a negative situation/statement and spin it into a positive, high-value frame.",
         "Broadens ifByXYouMeanY (which only reframes criticism into status)."),
        ("Connotation Swap", "Wk4 / 06 Denotation vs Connotation",
         "Restate the same fact with a more attractive/charming connotation ('player' -> 'outgoing social guy').",
         "New. Family with Reframe & Rephrase for Value."),
        ("Rephrase for Value", "Wk0 / Phrasing",
         "Rewrite a needy/low-value line into a higher-value, less agenda-driven one.",
         "New. Family with Reframe & Connotation Swap (could merge into one 'Phrasing' exercise)."),
        ("Storytelling (3-Act)", "Wk6 / 02 General Storytelling",
         "Given a prompt/random object, tell an engaging beginning-middle-end micro-story.",
         "New. Longer-form than current single-line drills."),
        ("Purposeful Story", "Wk6 / 03 Purposeful Storytelling",
         "Tell a (true) story chosen/shaped to convey a specific desirable trait.",
         "New. Variant of Storytelling."),
        ("Show Don't Tell", "Wk6 / 04 Objective Correlative",
         "Imply a trait through a concrete detail/event instead of claiming it.",
         "New. Advanced variant of Storytelling."),
        ("Qualification Tease", "Wk5 / 02 Tease",
         "Give a playful challenge that makes her qualify herself to you.",
         "Overlaps questionAnswerTease; could be a mode of it."),
        ("Deepening Question", "Wk8 / 13 Eliciting Values",
         "Turn a surface answer into an emotional 'why does that matter to you' question.",
         "Overlaps vibing (attunement) but narrower and more teachable."),
        ("Instant Narrative / Role", "Wk0 / Building a Narrative",
         "Assign a playful instant dynamic ('you're my runaway ex') and banter from inside it.",
         "New creative drill; cousin of yesAnd and a natural fit for the roleplay engine."),
    ]
    for i, (name, src, what, rel) in enumerate(new_ex, start=1):
        ws2.append([i, name, src, what, rel])

    for c in ws2[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    for i, w in enumerate([5, 26, 30, 72, 60], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    for row in ws2.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
    ws2.freeze_panes = "A2"

    # ---- Note: archetypes as roleplay fuel (small callout sheet) ----
    ws3 = wb.create_sheet("Roleplay Persona Fuel")
    ws3.append(["Archetype", "Source", "Why it matters for riffy"])
    arche = [
        ("The Princess", "Wk8 / Archetypes 1", "Bored/privileged; needs challenge & polarity. Maps to a roleplay character with a distinct shit-test style."),
        ("Smart-Good Girl", "Wk8 / Archetypes 2", "Values conformity; arrogance reads as inauthentic. Needs pacing-then-leading. Distinct character voice."),
        ("First-Generation American", "Wk8 / Archetypes 3", "Resilient, hard-working, wary of unearned success. Distinct values/backstory for a character."),
        ("The Entrepreneur", "Wk8 / Archetypes 4", "Ambitious; respects competence & 'working smart'. Distinct character voice."),
    ]
    for a in arche:
        ws3.append(list(a))
    for c in ws3[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    for i, w in enumerate([28, 22, 90], start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    for row in ws3.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
    ws3.freeze_panes = "A2"

    wb.save(OUT)
    total = len(rows)
    trues = sum(1 for r in rows if D[r["path"]][0])
    print(f"Wrote {OUT}")
    print(f"Rows: {total} | Can-form-exercise TRUE: {trues} | FALSE: {total - trues}")


if __name__ == "__main__":
    main()
