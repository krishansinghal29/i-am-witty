#!/usr/bin/env python3
"""Step 3 — compile final/exercise-research.xlsx from all pipeline outputs.

Sheets: Overview, Master Exercises, Providers, App Metadata, Provider Exercises,
Coverage Matrix. Run with the project's .venv python (openpyxl).
"""
import json, glob, os, re
from collections import defaultdict, OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz, process

FUZZY_CUTOFF = 88   # token_sort_ratio floor for a fuzzy matched_exercise_id

BASE = os.path.join(os.path.dirname(__file__), "..")
OUT = os.path.join(BASE, "final", "exercise-research.xlsx")
DOMAINS = ["storytelling", "humor-jokes", "improv", "quick-wit",
           "dating-social", "oratory", "conversation", "voice-tonality"]

def load(p):
    rows = []
    for f in sorted(glob.glob(os.path.join(BASE, p))):
        for line in open(f):
            line = line.strip()
            if line:
                try: rows.append(json.loads(line))
                except: pass
    return rows

def norm_name(s):
    s = (s or "").strip().lower().strip('"\'')
    s = re.sub(r"\s+", " ", s)
    return s.strip(" .!?:;-")

_FILLER = {"game","games","exercise","exercises","drill","drills","warmup","warmups",
           "technique","techniques","the","a","an","of","your","with"}
def norm_agg(s):
    """Aggressive normalization: drop punctuation, articles, filler words."""
    s = (s or "").lower().replace("&", "and")
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    toks = [t for t in re.sub(r"\s+", " ", s).strip().split() if t not in _FILLER]
    return " ".join(toks)

# ---- provider canonicalization (merge 2b + 2c variants of same provider) ----
ALIAS = {  # normalized-key overrides
    "charisma on command": "charisma university",
    "toastmasters international": "toastmasters",
    "toastmasters pathways": "toastmasters",
}
DISPLAY = {  # canonical-key -> preferred display name
    "charisma university": "Charisma University (Charisma on Command)",
    "toastmasters": "Toastmasters (International / Pathways)",
}
def prov_key(name):
    k = (name or "").lower()
    k = re.split(r"\s+[—–]\s+|\s+\(", k)[0]   # cut at em/en dash or " ("
    k = re.sub(r"\s+", " ", k).strip()
    return ALIAS.get(k, k)

_display_seen = {}
def prov_display(name):
    k = prov_key(name)
    if k in DISPLAY:
        return DISPLAY[k]
    # else longest original name seen for this key
    cur = _display_seen.get(k)
    if cur is None or len(name) > len(cur):
        _display_seen[k] = name
    return _display_seen[k]

# ---- styling helpers ----
HEAD_FILL = PatternFill("solid", fgColor="2F4F4F")
HEAD_FONT = Font(bold=True, color="FFFFFF")
def style_sheet(ws, widths, wrap_cols=()):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c in ws[1]:
        c.fill = HEAD_FILL; c.font = HEAD_FONT
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in wrap_cols:
        for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

def join(v):
    if isinstance(v, list):
        return ", ".join(str(x) for x in v)
    return "" if v is None else str(v)

# ================= load data =================
master = load("output/step1/master-exercises.jsonl")
providers = load("output/step2/providers.jsonl")
meta = load("output/step2/app-metadata/meta-*.jsonl")
pex = load("output/step2/provider-exercises/*.jsonl") + load("output/step2/coach-programs/*.jsonl")

# master lookups for matched_exercise_id (name + aliases):
#  - loose: norm_name string equality
#  - agg:   aggressive-normalized string equality
#  - fuzzy: rapidfuzz token_sort_ratio over the aggressive keys
name2id, agg2id = {}, {}
for r in master:
    for nm in [r.get("name")] + (r.get("aliases") or []):
        k = norm_name(nm)
        if k and k not in name2id:
            name2id[k] = r["id"]
        a = norm_agg(nm)
        if a and a not in agg2id:
            agg2id[a] = r["id"]
agg_keys = list(agg2id.keys())

def match_master(name):
    """Return (id, method, score) or ('', '', '')."""
    loose = norm_name(name); agg = norm_agg(name)
    if loose in name2id:
        return name2id[loose], "exact", 100
    if agg in agg2id:
        return agg2id[agg], "exact-norm", 100
    if len(agg) >= 5:
        m = process.extractOne(agg, agg_keys, scorer=fuzz.token_sort_ratio,
                               score_cutoff=FUZZY_CUTOFF)
        if m:
            return agg2id[m[0]], "fuzzy", round(m[1])
    return "", "", ""

# dedupe provider-exercises within canonical provider on normalized exercise name
seen = {}
pex_clean = []
for r in pex:
    cprov = prov_display(r.get("provider", ""))
    key = (prov_key(r.get("provider", "")), norm_name(r.get("exercise_name")))
    r = dict(r, _cprov=cprov)
    mid, mmethod, mscore = match_master(r.get("exercise_name"))
    r["matched_exercise_id"] = mid
    r["match_method"] = mmethod
    r["match_score"] = mscore
    if key in seen:
        # keep the row with the longer description
        if len(r.get("description", "")) > len(seen[key].get("description", "")):
            idx = seen[key]["_idx"]
            r["_idx"] = idx
            pex_clean[idx] = r; seen[key] = r
        continue
    r["_idx"] = len(pex_clean)
    seen[key] = r
    pex_clean.append(r)

# per-canonical-provider stats
prov_ex = defaultdict(list)
for r in pex_clean:
    prov_ex[r["_cprov"]].append(r)
meta_by_name = {m.get("provider"): m for m in meta}

# ================= workbook =================
wb = Workbook()

# ---- Overview ----
ws = wb.active; ws.title = "Overview"
matched = sum(1 for r in pex_clean if r["matched_exercise_id"])
m_exact = sum(1 for r in pex_clean if r.get("match_method","").startswith("exact"))
m_fuzzy = sum(1 for r in pex_clean if r.get("match_method") == "fuzzy")
ov = [
    ["i-am-witty / riffy — Exercise & Provider Research", ""],
    ["Generated", "2026-06-10"],
    ["", ""],
    ["Sheet", "Contents"],
    ["Master Exercises", f"{len(master)} unique exercises across 8 domains (Step 1, de-duplicated)"],
    ["Providers", f"{len(providers)} apps/courses/coaches/schools/communities (Step 2a)"],
    ["App Metadata", f"{len(meta)} app-type providers: funding, revenue, downloads, ownership (Step 2d)"],
    ["Provider Exercises", f"{len(pex_clean)} exercises taught by providers (Step 2b deep-dives + 2c dossiers), de-duped per provider"],
    ["  of which matched", f"{matched} linked back to a Master Exercise ({m_exact} exact/normalized name, {m_fuzzy} fuzzy >= {FUZZY_CUTOFF}). See Match Method / Match Score cols. Low rate is expected: the catalog and provider corpus were built independently, so most provider drills either aren't in the master set or use a lexically different name (name-matching only, no semantic match)."],
    ["Coverage Matrix", "providers x 8 domains, counts of exercises offered"],
    ["", ""],
    ["Domains", ", ".join(DOMAINS)],
    ["Note", "Provider names merged across Step 2b/2c (e.g. Greg Dean, Toastmasters); exercise dupes within a provider collapsed, keeping the most detailed description."],
]
for row in ov:
    ws.append(row)
ws["A1"].font = Font(bold=True, size=14)
for r in range(4, 11):
    ws[f"A{r}"].font = Font(bold=True)
ws.column_dimensions["A"].width = 24; ws.column_dimensions["B"].width = 95
for row in ws.iter_rows(min_row=2):
    row[1].alignment = Alignment(wrap_text=True, vertical="top")

# ---- Master Exercises ----
ws = wb.create_sheet("Master Exercises")
cols = ["id","name","aliases","domains","also_helps","description","instructions",
        "format","duration_minutes","difficulty","setup","skill_targets","origin",
        "source_urls","variations","n_source_domains"]
ws.append([c.replace("_"," ").title() for c in cols])
for r in master:
    ws.append([join(r.get(c, "")) for c in cols])
style_sheet(ws, [9,28,22,20,18,46,60,9,11,12,22,26,22,34,30,9], wrap_cols=(6,7,15))

# ---- Providers ----
ws = wb.create_sheet("Providers")
pcols = ["name","type","url","domains","pricing","offers_exercises","what_it_is",
         "has_structured_program","worth_deep_dive","notes"]
hdr = [c.replace("_"," ").title() for c in pcols] + ["Deep Dived","# Exercises Found","Founded","Funding"]
ws.append(hdr)
for r in providers:
    disp = prov_display(r.get("name",""))
    exs = prov_ex.get(disp, [])
    m = meta_by_name.get(r.get("name"), {})
    ws.append([join(r.get(c,"")) for c in pcols] +
              ["yes" if exs else "no", len(exs),
               m.get("founded","") or m.get("launched",""), m.get("funding_total","")])
style_sheet(ws, [30,15,34,20,16,15,46,16,13,40,11,15,10,16], wrap_cols=(7,10))

# ---- App Metadata ----
ws = wb.create_sheet("App Metadata")
mcols = ["provider","company","url","founded","launched","founders","ceo","hq",
         "funding_total","funding_rounds","last_valuation","revenue_estimate",
         "downloads_estimate","users_estimate","employees","pricing","platforms",
         "ownership","confidence_notes","source_urls"]
ws.append([c.replace("_"," ").title() for c in mcols])
for m in meta:
    row = []
    for c in mcols:
        v = m.get(c, "")
        if c == "funding_rounds" and isinstance(v, list):
            v = " | ".join(f"{x.get('stage','')} {x.get('amount','')} {x.get('date','')} "
                           f"[{', '.join(x.get('investors',[]) or [])}]".strip() for x in v)
        row.append(join(v))
    ws.append(row)
style_sheet(ws, [26,22,30,10,11,26,18,20,14,40,14,16,16,14,11,30,16,24,46,34],
            wrap_cols=(10,16,19,20))

# ---- Provider Exercises ----
ws = wb.create_sheet("Provider Exercises")
ws.append(["Provider","Exercise Name","Description","Domains","Format","Evidence",
           "Matched Exercise ID","Match Method","Match Score","Source URLs"])
for r in sorted(pex_clean, key=lambda x: (x["_cprov"].lower(), norm_name(x.get("exercise_name")))):
    ws.append([r["_cprov"], join(r.get("exercise_name")), join(r.get("description")),
               join(r.get("domains")), join(r.get("format")), join(r.get("evidence")),
               r.get("matched_exercise_id",""), r.get("match_method",""),
               r.get("match_score",""), join(r.get("source_urls"))])
style_sheet(ws, [30,30,54,20,16,11,16,13,11,38], wrap_cols=(3,10))

# ---- Coverage Matrix ----
ws = wb.create_sheet("Coverage Matrix")
ws.append(["Provider","Type"] + DOMAINS + ["TOTAL"])
ptype = {prov_display(p.get("name","")): p.get("type","") for p in providers}
rows = []
for disp, exs in prov_ex.items():
    dc = {d: 0 for d in DOMAINS}
    for e in exs:
        for d in (e.get("domains") or []):
            if d in dc: dc[d] += 1
    rows.append((disp, ptype.get(disp, ""), [dc[d] for d in DOMAINS], len(exs)))
rows.sort(key=lambda x: -x[3])
for disp, typ, counts, tot in rows:
    ws.append([disp, typ] + counts + [tot])
# totals row
ws.append(["— TOTAL —", ""] +
          [sum(r[2][i] for r in rows) for i in range(len(DOMAINS))] +
          [sum(r[3] for r in rows)])
style_sheet(ws, [34,15] + [13]*len(DOMAINS) + [9])
ws[f"A{ws.max_row}"].font = Font(bold=True)

os.makedirs(os.path.join(BASE, "final"), exist_ok=True)
wb.save(OUT)
print("wrote", os.path.relpath(OUT, BASE))
print(f"  Master Exercises : {len(master)}")
print(f"  Providers        : {len(providers)}")
print(f"  App Metadata     : {len(meta)}")
print(f"  Provider Exercises: {len(pex_clean)} ({matched} matched: {m_exact} exact/norm + {m_fuzzy} fuzzy>={FUZZY_CUTOFF})")
print(f"  Coverage Matrix  : {len(rows)} providers")
